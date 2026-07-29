"""Layer 0 of the document-ingestion design (systemdesign/13-document-ingestion.md):
an extension -> markdown dispatcher. Deterministic, offline, no LLM, no network.

Mirrors microsoft/markitdown's dispatch shape (MIT) but swaps its pdfminer PDF
path for pypdfium2 + a pypdf fallback -- see the design doc's Provenance table
for why (AGPL/quality tradeoffs of the alternatives).

`extract_text(path)` is the one public entry point every caller (file_read,
later doc_digest) should route through.
"""

from __future__ import annotations

from pathlib import Path

import mammoth
import openpyxl
import pypdfium2 as pdfium
from markdownify import markdownify
from pypdf import PdfReader

# ponytail: per-sheet row cap keeps a whole workbook cheap to page through the
# chat loop; raise it (or teach doc_digest to page sheets) if a real workflow
# needs full sheets returned in one call.
_XLSX_ROW_CAP = 200
# Same reasoning per sheet and per page. These bound the WORK, not just the
# output: extraction runs under asyncio.to_thread, which is not cancellable, so
# a 5000-page PDF would hold a pool thread for the process lifetime with no
# timeout able to fire. Callers cap the resulting text separately.
_XLSX_SHEET_CAP = 20
_PDF_PAGE_CAP = 100
# Hard refusal above this. Every converter below reads the whole file into
# memory, and file_read's caller is Tier 1 inside roots (no approval), so an
# unbounded read is an unattended OOM of the Brain.
_MAX_BYTES = 64 * 1024 * 1024


def _truncation_note(kind: str, cap: int, total: int) -> str:
    return f"\n\n... [truncated at {cap} {kind} of {total} total]"


def _extract_pdf(path: Path) -> str:
    text = ""
    total = 0
    try:
        pdf = pdfium.PdfDocument(str(path))
        try:
            total = len(pdf)
            parts = [
                pdf[i].get_textpage().get_text_range() for i in range(min(total, _PDF_PAGE_CAP))
            ]
        finally:
            pdf.close()
        text = "\n\n".join(parts)
    except Exception:
        text = ""  # fall through to the pypdf fallback below
    if not text.strip():
        try:
            reader = PdfReader(str(path))
            if reader.is_encrypted:
                # Distinct from a scanned page: the text may be right there,
                # locked. Say so, so the remedy (supply the password) is honest.
                raise ValueError(f"{path.name} is an encrypted/password-protected PDF")
            total = len(reader.pages)
            text = "\n\n".join(
                reader.pages[i].extract_text() or "" for i in range(min(total, _PDF_PAGE_CAP))
            )
        except ValueError:
            raise
        except Exception:
            text = ""
    if text.strip() and total > _PDF_PAGE_CAP:
        text += _truncation_note("pages", _PDF_PAGE_CAP, total)
    if not text.strip():
        raise ValueError(
            f"no extractable text in {path.name} (scanned PDF? Layer 0 is text-only, no OCR yet)"
        )
    return text


def _extract_docx(path: Path) -> str:
    # mammoth's own Markdown writer is deprecated upstream ("generating HTML and
    # using a separate library to convert the HTML to Markdown is recommended");
    # markdownify is already the .html path's converter, so reuse it.
    with path.open("rb") as handle:
        result = mammoth.convert_to_html(handle)
    return markdownify(result.value)


def _fmt_row(cells: tuple, width: int) -> str:
    values = [("" if c is None else str(c)) for c in cells]
    values += [""] * (width - len(values))
    return "| " + " | ".join(values) + " |"


def _extract_xlsx(path: Path) -> str:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sections = []
        for name in wb.sheetnames[:_XLSX_SHEET_CAP]:
            ws = wb[name]
            rows = []
            truncated = False
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= _XLSX_ROW_CAP:
                    truncated = True
                    break
                rows.append(row)
            if not rows:
                sections.append(f"## {name}\n\n(empty sheet)")
                continue
            width = len(rows[0])
            lines = [_fmt_row(rows[0], width), "| " + " | ".join(["---"] * width) + " |"]
            lines += [_fmt_row(r, width) for r in rows[1:]]
            body = "\n".join(lines)
            if truncated:
                body += f"\n\n... [truncated at {_XLSX_ROW_CAP} rows of sheet '{name}'; more rows exist]"
            sections.append(f"## {name}\n\n{body}")
        if not sections:
            return "(empty workbook)"
        out = "\n\n".join(sections)
        if len(wb.sheetnames) > _XLSX_SHEET_CAP:
            out += _truncation_note("sheets", _XLSX_SHEET_CAP, len(wb.sheetnames))
        return out
    finally:
        wb.close()


def _extract_html(path: Path) -> str:
    html = path.read_text(encoding="utf-8", errors="replace")
    return markdownify(html)


_CONVERTERS = {
    ".pdf": _extract_pdf,
    ".docx": _extract_docx,
    ".xlsx": _extract_xlsx,
    ".html": _extract_html,
    ".htm": _extract_html,
}


def extract_text(path: Path) -> str:
    """Return markdown (or plain text) for `path`, deterministically and
    offline. `.md`/`.txt`/source-code files and anything else that decodes as
    UTF-8 text pass through unchanged -- no converter list to maintain for
    every language extension. Raises ValueError naming the format when there
    is genuinely nothing extractable: a binary format with no converter above,
    or (PDF-specific) a scanned page with no text layer at all -- or, before any
    of that, when the file is too large to read into memory at all.
    """
    size = path.stat().st_size
    if size > _MAX_BYTES:
        raise ValueError(
            f"{path.name} is {size // (1024 * 1024)}MB; refusing to read more than "
            f"{_MAX_BYTES // (1024 * 1024)}MB into memory"
        )
    ext = path.suffix.lower()
    converter = _CONVERTERS.get(ext)
    if converter is not None:
        return converter(path)
    data = path.read_bytes()
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        raise ValueError(
            f"no extractable text: {ext or '(no extension)'} is not a supported text or document format"
        ) from None
