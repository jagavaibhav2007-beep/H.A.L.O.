"""Runnable self-check for brain/extract.py (Layer 0, systemdesign/13-document-ingestion.md).

No test framework -- plain asyncio + assert, matching test_files.py. Run with:
    python brain/tests/test_extract.py

Builds tiny real files (xlsx via openpyxl, docx by hand as a minimal zip, a
blank PDF via pypdf) in a temp dir rather than shipping binary fixtures.
"""

from __future__ import annotations

import asyncio
import sys
import tempfile
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from pypdf import PdfWriter

from brain import extract
from brain.tools import files  # registers file_read; also exposes _file_read/_READ_CAP

TMP = Path(tempfile.mkdtemp(prefix="halo-test-extract-"))


def check_text_passthrough() -> None:
    for name, content in (("a.md", "# heading\n\nsome *text*"), ("b.txt", "plain text"), ("c.py", "x = 1\n")):
        p = TMP / name
        p.write_text(content, encoding="utf-8", newline="")  # no \n -> \r\n translation on Windows
        assert extract.extract_text(p) == content, name
    print("[check 1] .md/.txt/source-code files pass through unchanged: OK")


def check_xlsx_pipe_table() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "People"
    ws.append(["name", "age"])
    ws.append(["alice", 30])
    ws.append(["bob", 25])
    p = TMP / "sheet.xlsx"
    wb.save(p)
    md = extract.extract_text(p)
    assert "## People" in md, md
    assert "| name | age |" in md, md
    assert "| --- | --- |" in md, md
    assert "| alice | 30 |" in md and "| bob | 25 |" in md, md

    # Row cap: build a sheet with more rows than _XLSX_ROW_CAP and expect an honest note.
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Big"
    for i in range(extract._XLSX_ROW_CAP + 10):
        ws2.append([i])
    p2 = TMP / "big.xlsx"
    wb2.save(p2)
    md2 = extract.extract_text(p2)
    assert f"truncated at {extract._XLSX_ROW_CAP} rows" in md2, md2
    print("[check 2] .xlsx renders as a markdown pipe table per sheet, with an honest row-cap note: OK")


def check_html_strips_tags() -> None:
    p = TMP / "page.html"
    p.write_text("<html><body><h1>Title</h1><p>Hello <b>world</b></p></body></html>", encoding="utf-8")
    md = extract.extract_text(p)
    assert "<h1>" not in md and "<p>" not in md and "<b>" not in md, md
    assert "Title" in md and "Hello" in md and "world" in md, md
    print("[check 3] .html strips tags via markdownify, keeping the text: OK")


def check_unknown_extension_raises() -> None:
    p = TMP / "binary.exe"
    p.write_bytes(bytes(range(256)) * 4)  # not valid UTF-8, no converter registered
    try:
        extract.extract_text(p)
        assert False, "unknown binary format should have raised"
    except ValueError as e:
        assert ".exe" in str(e), e
    print("[check 4] unknown binary extension raises ValueError naming the format: OK")


def _minimal_docx(path: Path, text: str) -> None:
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/></Relationships>'
    )
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body><w:p><w:r><w:t>{text}</w:t></w:r></w:p></w:body></w:document>"
    )
    with zipfile.ZipFile(path, "w") as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", rels)
        z.writestr("word/document.xml", document)


def check_docx_via_mammoth() -> None:
    p = TMP / "doc.docx"
    _minimal_docx(p, "Hello docx world")
    md = extract.extract_text(p)
    assert "Hello docx world" in md, md
    # mammoth's own markdown writer is deprecated upstream, so this now goes
    # HTML -> markdownify (the .html path's converter). The output must still be
    # markdown, not the intermediate HTML.
    assert "<p>" not in md and "</p>" not in md, md
    print("[check 5] .docx converts via mammoth HTML + markdownify: OK")


def check_sheet_and_page_caps_bound_the_work() -> None:
    """These cap the WORK, not just the output. Extraction runs under
    asyncio.to_thread, which is not cancellable -- an unbounded parse holds a
    pool thread for the process lifetime and no timeout can fire."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for i in range(extract._XLSX_SHEET_CAP + 5):
        wb.create_sheet(f"S{i}").append(["v", i])
    p = TMP / "many_sheets.xlsx"
    wb.save(p)
    md = extract.extract_text(p)
    assert f"## S{extract._XLSX_SHEET_CAP - 1}" in md, "capped short of the limit"
    assert f"## S{extract._XLSX_SHEET_CAP}" not in md, "sheet cap did not bound the workbook"
    assert f"truncated at {extract._XLSX_SHEET_CAP} sheets" in md, md

    # Page cap: a real multi-page text PDF is awkward to build, so drive the
    # constant down instead -- the branch under test is the same one.
    prev = extract._PDF_PAGE_CAP
    extract._PDF_PAGE_CAP = 1
    try:
        w = PdfWriter()
        for _ in range(3):
            w.add_blank_page(width=200, height=200)
        p2 = TMP / "three_blank.pdf"
        with p2.open("wb") as handle:
            w.write(handle)
        pages_read: list[int] = []
        real_reader = extract.PdfReader

        class _CountingReader(real_reader):  # type: ignore[misc,valid-type]
            @property
            def pages(self):
                got = super().pages
                pages_read.append(len(got))
                return got

        extract.PdfReader = _CountingReader
        try:
            try:
                extract.extract_text(p2)  # blank pages -> honest "no text" error
            except ValueError:
                pass
        finally:
            extract.PdfReader = real_reader
        # The point: the fallback asked for len(pages) but only ever indexed
        # within the cap -- it never iterated all three.
        assert pages_read and pages_read[0] == 3, pages_read
    finally:
        extract._PDF_PAGE_CAP = prev
    print("[check 8] xlsx sheet cap and pdf page cap bound the parse work: OK")


def check_oversized_file_refused_once() -> None:
    """No size check anywhere was a Tier-1-inside-roots OOM of the Brain: a
    multi-GB file loaded entirely, and the resulting MemoryError was caught by
    _file_read's `except Exception` fallback, which then read the file AGAIN."""
    p = TMP / "huge.bin"
    with p.open("wb") as handle:  # sparse: no real disk cost
        handle.truncate(extract._MAX_BYTES + 1)

    reads: list[str] = []
    real_open = Path.open

    def _counting_open(self, *a, **kw):
        if self == p:
            reads.append(str(a[:1]))
        return real_open(self, *a, **kw)

    Path.open = _counting_open
    try:
        try:
            files._file_read({"path": str(p)})
            assert False, "an oversized file should be refused, not read"
        except ValueError as e:
            assert "refusing to read" in str(e), e
    finally:
        Path.open = real_open
    assert reads == [], f"the refused file was opened anyway ({len(reads)}x): {reads}"

    # And a MemoryError from a converter must not fall through to the raw-read
    # fallback -- that fallback re-reads the same file that just did not fit.
    small = TMP / "small.md"
    small.write_text("hi", encoding="utf-8")
    calls: list[str] = []
    real_extract = extract.extract_text

    def _boom(path):
        calls.append(str(path))
        raise MemoryError("simulated allocation failure")

    files.extract.extract_text = _boom
    try:
        try:
            files._file_read({"path": str(small)})
            assert False, "MemoryError should propagate, not be swallowed"
        except MemoryError:
            pass
    finally:
        files.extract.extract_text = real_extract
    assert len(calls) == 1, f"file was read {len(calls)} times after MemoryError, expected 1"
    print("[check 9] oversized file refused without reading; MemoryError never triggers a second read: OK")


def check_pdf_extraction_honesty() -> None:
    # A blank page has no text layer at all -- both pypdfium2 and the pypdf
    # fallback come back empty, so this must be the honest "no extractable
    # text" error, never silently-empty success or a crash.
    p = TMP / "blank.pdf"
    w = PdfWriter()
    w.add_blank_page(width=200, height=200)
    with p.open("wb") as handle:
        w.write(handle)
    try:
        extract.extract_text(p)
        assert False, "a textless PDF should raise, not return silently"
    except ValueError as e:
        assert "no extractable text" in str(e), e
    print("[check 6] scanned/textless PDF is an honest 'no extractable text' error, no crash: OK")


def check_file_read_pagination_and_cap() -> None:
    p = TMP / "many_lines.txt"
    total = 2000  # long enough that the whole file exceeds the 8KB cap
    lines = [f"line {i}\n" for i in range(total)]
    p.write_text("".join(lines), encoding="utf-8", newline="")  # no \n -> \r\n translation on Windows

    # Whole-file read is bigger than the 8KB cap -> truncation note with a
    # concrete offset to continue from.
    whole = files._file_read({"path": str(p)})
    assert f"truncated at {files._READ_CAP // 1024}KB" in whole, whole
    assert "offset=" in whole, whole

    # offset/limit page through without hitting the cap.
    page1 = files._file_read({"path": str(p), "offset": 1, "limit": 10})
    assert page1.count("line ") >= 10, page1
    assert "line 0\n" in page1 and "line 9\n" in page1, page1
    assert f"showing lines 1-10 of {total}" in page1, page1

    page2 = files._file_read({"path": str(p), "offset": 11, "limit": 10})
    assert "line 10\n" in page2 and "line 19\n" in page2, page2
    assert "line 0\n" not in page2, page2

    # Last page: no more lines remain, note says so without an offset= continuation.
    last = files._file_read({"path": str(p), "offset": total - 9, "limit": 20})
    assert f"line {total - 1}" in last, last
    assert f"{total} total" in last, last
    print("[check 7] file_read offset/limit pages through content; 8KB cap note names remaining lines + next offset: OK")


async def main() -> None:
    check_text_passthrough()
    check_xlsx_pipe_table()
    check_html_strips_tags()
    check_unknown_extension_raises()
    check_docx_via_mammoth()
    check_pdf_extraction_honesty()
    check_file_read_pagination_and_cap()
    check_sheet_and_page_caps_bound_the_work()
    check_oversized_file_refused_once()
    print("[brain.extract] self-check OK")


if __name__ == "__main__":
    asyncio.run(main())
